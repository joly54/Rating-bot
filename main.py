try:
    import telebot
    import json
    import gspread
    from gspread.cell import Cell

    gc = gspread.service_account(filename="key.json")
    import datetime
    from openpyxl.utils import get_column_letter
    from gspread_formatting import *
    import time
    import threading
    import sys
    count = 16
    # Open a spreadsheet by title
    sh = gc.open("Rating")
    wk = sh.sheet1

    bot = telebot.TeleBot(#)
    try:
        with open("students.json", "r") as file:
            students = json.load(file)
            i = 1
            for student in students:
                print(
                    (
                        "{}) {}\t{}\t{}\t{}\n".format(
                            i,
                            students[student][0],
                            students[student][2],
                            student,
                            students[student][1],
                        )
                    )
                )
                i += 1
    except:
        students = {}
    try:
        with open("count.json", "r") as file:
            count = json.load(file)
            print(f"count: {count}")
    except:
        pass

    is_run=1
    date=""
    i=1
    def running():
        global backup_time
        backup_time = 0
        i=1
        while is_run==1:
            global date
            now = datetime.datetime.now()
            date_time = now.strftime("Time %H-%M-%S Date %d-%m-%y")
            date = now.strftime("%d-%m-%y")
            f=open(f"runing date: {date}.txt", "a")
            f.write(f"Working in {date_time}\n")
            f.close()
            i+=1
            if is_run==0:
                exit(1)
            if (time.time()-backup_time)>7200:
                f = open("students.json","rb")
                bot.send_document(800918003, f)
                backup_time=time.time()
            time.sleep(300)
    (threading.Thread(target=running)).start()

    def create(students):
        sorted_students = sorted(students.items(), key=lambda x: x[1][2])
        return [student[1] for student in sorted_students]
    def generate_color_array(steps):
                colors = []
                for i in range(steps):
                    green = 1 - (i / steps) * 0.8
                    blue = 0
                    red = i / steps
                    colors.append([red, green, blue])
                return colors
    def update(students):
        try:
            cells = []
            #wk.clear()
            cells.append(Cell(row=1, col=1, value="Position"))
            cells.append(Cell(row=1, col=2, value="Group"))
            cells.append(Cell(row=1, col=3, value="Name"))
            cells.append(Cell(row=1, col=4, value="Points"))
            cells.append(Cell(row=1, col=5, value="Telegram name"))
            #setup()
            for i in range(0, 50):
                if i< len(students):
                    cells.append(Cell(i + 2, 1, value=i + 1))
                    cells.append(Cell(i + 2, 2, students[i][0]))
                    cells.append(Cell(i + 2, 3, students[i][1]))
                    cells.append(Cell(i + 2, 4, students[i][2]))
                    cells.append(Cell(i + 2, 5, students[i][4]))
                else:
                    cells.append(Cell(i + 2, 1, value=""))
                    cells.append(Cell(i + 2, 2, value=""))
                    cells.append(Cell(i + 2, 3, value=""))
                    cells.append(Cell(i + 2, 4, value=""))
                    cells.append(Cell(i + 2, 5, value=""))
            formating=[]
            colors = generate_color_array((count+1)-round(count*0.15))
            fmt = cellFormat(
                backgroundColor=color(1, 1, 1),
            )
            formating.append((f'{get_column_letter(1)}{1}:{get_column_letter(500)}{500}', fmt))
            for i in range(1+round(count*0.15),count+1):
                    fmt = cellFormat(
                        backgroundColor=color(colors[i-round(count*0.15)][0], colors[i-round(count*0.15)][1], colors[i-round(count*0.15)][2]),
                        textFormat=textFormat(bold=False),
                        horizontalAlignment='CENTER'
                    )
                    formating.append((f'{get_column_letter(1)}{i+1}:{get_column_letter(5)}{i+1}', fmt))
            for i in range(1,+round(count*0.15)+1):
                    fmt = cellFormat(
                        backgroundColor=color(1, 0.875, 0),
                        textFormat=textFormat(bold=False),
                        horizontalAlignment='CENTER'
                    )
                    formating.append((f'{get_column_letter(1)}{i+1}:{get_column_letter(5)}{i+1}', fmt))
            fmt = cellFormat(
                backgroundColor=color(1, 1, 1),
                textFormat=textFormat(bold=True),
                horizontalAlignment='CENTER'
            )
            formating.append((f'{get_column_letter(1)}{1}:{get_column_letter(5)}{1}', fmt))
            if len(students)>count:
                for i in range(count, len(students)):
                    fmt = cellFormat(
                        backgroundColor=color(0.85, 0.85, 0.85),
                        textFormat=textFormat(bold=False),
                        horizontalAlignment='CENTER'
                    )
                    formating.append((f'{get_column_letter(1)}{i+2}:{get_column_letter(5)}{i+2}', fmt))
            fmt = cellFormat(
                        textFormat=textFormat(bold=True),
                        horizontalAlignment='CENTER'
                    )
            formating.append((f'{get_column_letter(1)}{1}:{get_column_letter(1)}{50}', fmt))
            format_cell_ranges(wk, formating)
            wk.update_cells(cells)
        except Exception as e:
            print(e)
            bot.send_message(800918003, str(e))


    update(create(students)[::-1])
    @bot.message_handler(content_types=["text"])
    def get_text_messages(message):
        text = message.text
        try:
            name = str(message.from_user.first_name)
            if message.from_user.last_name:
                name = name + " " + str(message.from_user.last_name)
            user_id = str(message.from_user.id)
            if not user_id in students:
                students[user_id] = ["", "", 0, "Waiting for group", name]
                bot.send_message(user_id, "Відправте номер групи\n1 для 1-AKIT\n2 для 2-AKIT")
            else:
                if message.text.startswith("/"):
                    command = message.text[1:]
                    if command == "edit":
                        bot.send_message(
                            user_id, "Відправте номер групи\n1 для 1-AKIT\n2 для 2-AKIT"
                        )
                        students[user_id][3] = "Waiting for group"
                    elif user_id == "800918003" and command.find("delete") != -1:
                        try:
                            del students[command[command.find(" ") + 1 :]]
                            with open("students.json", "w") as file:
                                json.dump(students, file)
                            bot.send_message(user_id, "Deleted")
                            update(create(students)[::-1])
                            bot.send_message(user_id, "Updated")
                        except Exception as e:
                            print(e)
                            bot.send_message(800918003, str(e))
                    elif user_id == "800918003" and command.find("get") != -1:
                        try:
                            bot.send_message(
                                800918003, str(students[command[command.find(" ") + 1 :]])
                            )
                        except Exception as e:
                            print(e)
                            bot.send_message(800918003, str(e))
                    elif user_id == "800918003" and command.find("list") != -1:
                        try:
                            res = ""
                            i = 1
                            bot.send_message(user_id, "Preparing...")
                            for student in students:
                                res = res + (
                                    "{}) Group: {}\nName: {}\nPoints: {}\nID: <code>{}</code>\nTG name: {}\nStatus: {}\n".format(
                                        i,
                                        students[student][0],
                                        students[student][1],
                                        students[student][2],
                                        student,
                                        students[student][4],
                                        students[student][3],
                                    )
                                )
                                username = str((bot.get_chat(student)).username)
                                if  username != "None":
                                    res = (
                                        res
                                        + f"Username: @{username}\n"
                                    )
                                res = res + "\n"
                                i += 1
                            bot.send_message(user_id, res, parse_mode="HTML")
                        except Exception as e:
                            print(e)
                            bot.send_message(800918003, str(e))
                    elif command.find("link") != -1:
                        try:
                            bot.send_message(
                                user_id,
                                "Link to rating: https://docs.google.com/spreadsheets/d/1p53_Pb_9qb_TUOZD7L0W-G29INtR8w32uVFAD1qDOwc/edit#gid=0",
                                disable_web_page_preview=True,
                            )
                        except Exception as e:
                            print(e)
                            bot.send_message(800918003, str(e))
                    elif user_id == "800918003" and command.find("update") != -1:
                        try:
                            update(create(students)[::-1])
                            bot.send_message(user_id, "Updated")
                        except Exception as e:
                            print(e)
                            bot.send_message(800918003, str(e))
                    elif user_id == "800918003" and command.find("set") != -1:
                        try:
                            global count
                            count = int(command[command.find(" ") + 1 :])
                            with open("count.json", "w") as file:
                                json.dump(count, file)
                            update(create(students)[::-1])
                            bot.send_message(user_id, "Updated")
                        except Exception as e:
                            print(e)
                            bot.send_message(800918003, str(e))
                    elif user_id == "800918003" and command.find("send_file") != -1:
                        try:
                            f = open(command[command.find(" ") + 1 :], "rb")
                            bot.send_document(message.chat.id, f)
                        except Exception as e:
                            print(e)
                            bot.send_message(800918003, str(e))
                    elif user_id == "800918003" and command.find("edit_user") != -1:
                        try:
                            new_data = (command[command.find(" ") + 1 :]).split("; ")
                            print(new_data)
                            # del students[new_data[0]]
                            if len(new_data) == 5:
                                students[str(new_data[0])] = [
                                    new_data[1],
                                    new_data[2],
                                    int(new_data[3]),
                                    "Registration finished",
                                    new_data[4],
                                ]
                                update(create(students)[::-1])
                                bot.send_message(user_id, "Updated")
                            elif len(new_data) == 4:
                                students[str(new_data[0])] = [
                                    new_data[1],
                                    new_data[2],
                                    int(new_data[3]),
                                    "Registration finished",
                                    students[str(new_data[0])][4],
                                ]
                                update(create(students)[::-1])
                                bot.send_message(user_id, "Updated")
                            else:
                                bot.send_message(user_id, "Bad data")
                        except Exception as e:
                            print(e)
                            bot.send_message(800918003, str(e))
                    elif user_id == "800918003" and command.find("run") != -1:
                        try:
                            print(
                                "Code:\n\n{}\n\n--------------------------\n\nResult:\n".format(
                                    command[command.find(" ") + 1 :]
                                )
                            )
                            exec(command[command.find(" ") + 1 :])
                        except Exception as e:
                            print(e)
                            bot.send_message(800918003, str(e))
                    elif user_id == "800918003" and command.find("logs") != -1:
                        try:
                            f = open(f"log: {date}.txt")
                            bot.send_document(message.chat.id, f)
                        except Exception as e:
                            print(e)
                            bot.send_message(800918003, str(e))
                    elif user_id != "800918003" and (
                        command.find("delete") != -1
                        or command.find("get") != -1
                        or command.find("list") != -1
                        or command.find("link") != -1
                        or command.find("update") != -1
                        or command.find("set") != -1
                        or command.find("send_file") != -1
                        or command.find("edit_user") != -1
                        or command.find("run") != -1
                        or command.find("logs") != -1
                        ):
                        bot.send_message(user_id, "Access denied, aviable only for admin(@joly541)")
                elif students[user_id][3] == "Waiting for group":
                    if text.isnumeric() and (int(text) == 1 or int(text) == 2):
                        students[user_id][0] = text + "-AKIT"
                        bot.send_message(
                            user_id, "Вашу групу додано, тепер надішліть своє ім'я та прізвище"
                        )
                        students[user_id][3] = "Waiting for name"
                    else:
                        bot.send_message(
                            user_id, "Не вірні данні\nВідправте номер групи\n1 для 1-AKIT\n2 для 2-AKIT"
                        )

                elif students[user_id][3] == "Waiting for name":
                    students[user_id][1] = text
                    students[user_id][3] = "Waiting for points"
                    bot.send_message(
                        user_id,
                        "Ваше ім’я додано, тепер надішліть свої бали (без історії та англійської) цілим числом",
                    )
                elif students[user_id][3] == "Waiting for points":
                    if text.isnumeric():
                        students[user_id][2] = int(text)
                        bot.send_message(user_id, "Ваші бали додано")
                        students[user_id][3] = "Registration finished"
                        update(create(students)[::-1])
                        bot.send_message(
                            user_id,
                            "Посилання на рейтинг: https://docs.google.com/spreadsheets/d/1p53_Pb_9qb_TUOZD7L0W-G29INtR8w32uVFAD1qDOwc/edit#gid=0",
                            disable_web_page_preview=True,
                        )
                    else:
                        bot.send_message(user_id, "Невірні данні\nнадішліть свої бали (без історії та англійської) цілим числом")
                elif (
                    students[user_id][3] == "Registration finished"
                    and user_id != "800918003"
                ):
                    bot.send_message(
                        user_id,
                        "Посилання на рейтинг: https://docs.google.com/spreadsheets/d/1p53_Pb_9qb_TUOZD7L0W-G29INtR8w32uVFAD1qDOwc/edit#gid=0",
                        disable_web_page_preview=True,
                    )
            try:
                if user_id != "800918003":
                    now = datetime.datetime.now()
                    # Use the datetime.strftime method to specify the format of the date and time
                    date_time = now.strftime("Time %H-%M-%S Date %d-%m-%y")
                    print(
                        f"New message Name: {name}\tUser id:{user_id}\t{date_time}\tStatus: {students[user_id][3]}\tText: {text}\n"
                    )
                    f = open(f"log: {date}.txt", "a")
                    f.write(
                        f"New message Name: {name}\tUser id:{user_id}\t{date_time}\tStatus: {students[user_id][3]}\tText: {text}\n"
                    )
                    f.close
            except Exception as e:
                print(e)
                bot.send_message(800918003, str(e))
            with open("students.json", "w") as file:
                json.dump(students, file)
        except Exception as e:
            print(e)
            bot.send_message(800918003, str(e))


    bot.polling(none_stop=True, interval=0)
except Exception as e:
            print(e)
            f = open("Fatal error.txt", "a")
            f.write(
            str(e)
            )
            f.close
            try:
                bot.send_message(800918003, str(e))
            except:
                pass
            bot.polling(none_stop=True, interval=0)
